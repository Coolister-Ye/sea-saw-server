"""Generate Purchase Contract Excel file from a PurchaseOrder instance and return BytesIO."""

import os
from io import BytesIO

from openpyxl import load_workbook

import sea_saw_export
from sea_saw_export.xlsx import fix_copied_sheet, PRODUCT_FIRST_ROW, setup_product_rows, fill_product_row, fill_header
from sea_saw_export.xlsx.config import DocConfig

TEMPLATE_PATH = os.path.join(
    os.path.dirname(sea_saw_export.__file__), "templates", "PC_template.xlsx"
)

PC_CONFIG = DocConfig(
    template_sheet="TEMPLATE",
    doc_title="PURCHASE CONTRACT",
    stamp_col_shift=0,
    template_slots=1,
    total_row_offset=24,
    kg_per_carton=20,
    print_area_rows=38,
    buyer_sig_row_base=34,
    bank_details_row_base=29,
    title_party="Buyer",
)


def generate_pc_xlsx(purchase_order) -> BytesIO:
    """
    Generate a Purchase Contract XLSX for the given PurchaseOrder instance.
    Returns a BytesIO object ready to stream as a file response.
    """
    from .adapter import purchase_order_to_pc_data

    header, products = purchase_order_to_pc_data(purchase_order)

    wb = load_workbook(TEMPLATE_PATH, rich_text=True)
    template_ws = wb[PC_CONFIG.template_sheet]

    ws = wb.copy_worksheet(template_ws)
    fix_copied_sheet(template_ws, ws, PC_CONFIG)
    ws.title = f"PC-{purchase_order.purchase_code}"[:31]

    for name in list(wb.sheetnames):
        if name != ws.title:
            del wb[name]

    n_products = max(len(products), 1)
    total_row = setup_product_rows(ws, n_products, PC_CONFIG)

    for i, product in enumerate(products):
        fill_product_row(ws, PRODUCT_FIRST_ROW + i, product, PC_CONFIG)

    fill_header(ws, header, total_row, PC_CONFIG)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pc_bulk_xlsx(purchase_orders) -> BytesIO:
    """
    Generate a single XLSX with one sheet per purchase order, sorted by created_at ascending.
    Returns a BytesIO object ready to stream as a file response.
    """
    from .adapter import purchase_order_to_pc_data

    sorted_orders = sorted(purchase_orders, key=lambda po: po.created_at)

    wb = load_workbook(TEMPLATE_PATH, rich_text=True)
    template_ws = wb[PC_CONFIG.template_sheet]

    for po in sorted_orders:
        header, products = purchase_order_to_pc_data(po)
        ws = wb.copy_worksheet(template_ws)
        fix_copied_sheet(template_ws, ws, PC_CONFIG)
        ws.title = f"PC-{po.purchase_code}"[:31]

        n_products = max(len(products), 1)
        total_row = setup_product_rows(ws, n_products, PC_CONFIG)
        for i, product in enumerate(products):
            fill_product_row(ws, PRODUCT_FIRST_ROW + i, product, PC_CONFIG)
        fill_header(ws, header, total_row, PC_CONFIG)

    del wb[PC_CONFIG.template_sheet]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
