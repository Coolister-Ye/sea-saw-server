"""Generate PI Excel file from an Order instance and return BytesIO."""

import os
from io import BytesIO

from openpyxl import load_workbook

import sea_saw_export
from sea_saw_export.xlsx import fix_copied_sheet, PRODUCT_FIRST_ROW, setup_product_rows, fill_product_row, fill_header
from sea_saw_export.xlsx.config import DocConfig

TEMPLATE_PATH = os.path.join(
    os.path.dirname(sea_saw_export.__file__), "templates", "PI_template.xlsx"
)

PI_CONFIG = DocConfig(
    template_sheet="TEMPLATE",
    doc_title="PROFORMA INVOICE",
    stamp_col_shift=0,
    template_slots=1,
    total_row_offset=24,
    kg_per_carton=20,
    print_area_rows=38,
    buyer_sig_row_base=34,
    bank_details_row_base=29,
)


def generate_pi_bulk_xlsx(orders) -> BytesIO:
    """
    Generate a single XLSX with one sheet per order.
    Returns a BytesIO object ready to stream as a file response.
    """
    from .adapter import order_to_pi_data

    wb = load_workbook(TEMPLATE_PATH, rich_text=True)
    template_ws = wb[PI_CONFIG.template_sheet]

    for order in orders:
        header, products = order_to_pi_data(order)
        ws = wb.copy_worksheet(template_ws)
        fix_copied_sheet(template_ws, ws, PI_CONFIG)
        ws.title = f"PI-{order.order_code}"[:31]

        n_products = max(len(products), 1)
        total_row = setup_product_rows(ws, n_products, PI_CONFIG)
        for i, product in enumerate(products):
            fill_product_row(ws, PRODUCT_FIRST_ROW + i, product, PI_CONFIG)
        fill_header(ws, header, total_row, PI_CONFIG)

    del wb[PI_CONFIG.template_sheet]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pi_xlsx(order) -> BytesIO:
    """
    Generate a Proforma Invoice XLSX for the given Order instance.
    Returns a BytesIO object ready to stream as a file response.
    """
    from .adapter import order_to_pi_data

    header, products = order_to_pi_data(order)

    wb = load_workbook(TEMPLATE_PATH, rich_text=True)
    template_ws = wb[PI_CONFIG.template_sheet]

    ws = wb.copy_worksheet(template_ws)
    fix_copied_sheet(template_ws, ws, PI_CONFIG)
    ws.title = f"PI-{order.order_code}"[:31]

    for name in list(wb.sheetnames):
        if name != ws.title:
            del wb[name]

    n_products = max(len(products), 1)
    total_row = setup_product_rows(ws, n_products, PI_CONFIG)

    for i, product in enumerate(products):
        fill_product_row(ws, PRODUCT_FIRST_ROW + i, product, PI_CONFIG)

    fill_header(ws, header, total_row, PI_CONFIG)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
