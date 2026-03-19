"""Fill product rows and header fields into a worksheet (copied from order_files/src/writer.py)."""

from copy import copy

from openpyxl.styles import Font, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from .config import PRODUCT_FIRST_ROW
from .utils import parse_date, specs_to_multiline, safe_str


def _write_cell(ws, row, col, value):
    """Write value to a cell, resolving merged ranges to their top-left corner."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                cell = ws.cell(row=rng.min_row, column=rng.min_col)
                break
    cell.value = value


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def shift_rows_down(ws, from_row, shift_by):
    """Move all rows >= from_row downward by shift_by rows."""
    if shift_by <= 0:
        return

    to_shift = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= from_row:
            to_shift.append((rng.min_col, rng.min_row, rng.max_col, rng.max_row))
            ws.unmerge_cells(str(rng))

    for r in range(ws.max_row, from_row - 1, -1):
        for c in range(1, ws.max_column + 1):
            src = ws.cell(row=r, column=c)
            dst = ws.cell(row=r + shift_by, column=c)
            dst.value = src.value
            _copy_cell_style(src, dst)
            src.value = None
            src.font = Font()
            src.border = Border()
            src.fill = PatternFill()
        ws.row_dimensions[r + shift_by].height = ws.row_dimensions[r].height

    for min_col, min_row, max_col, max_row in to_shift:
        ws.merge_cells(
            f"{get_column_letter(min_col)}{min_row + shift_by}"
            f":{get_column_letter(max_col)}{max_row + shift_by}"
        )


def setup_product_rows(ws, n_products, cfg):
    """
    Ensure ws has n_products product rows starting at PRODUCT_FIRST_ROW.
    Returns the actual Total row number after expansion.
    """
    if n_products > 1:
        shift_rows_down(ws, PRODUCT_FIRST_ROW + 1, n_products - 1)

    ref_row = PRODUCT_FIRST_ROW
    for i in range(1, n_products):
        r = PRODUCT_FIRST_ROW + i
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= r <= rng.max_row:
                ws.unmerge_cells(str(rng))
        for c in range(2, 10):
            new_cell = ws.cell(row=r, column=c)
            new_cell.value = None
            _copy_cell_style(ws.cell(row=ref_row, column=c), new_cell)
        ws.merge_cells(f"B{r}:C{r}")
        ws.merge_cells(f"D{r}:E{r}")
        ws.row_dimensions[r].height = ws.row_dimensions[ref_row].height

    return cfg.total_row_offset + (n_products - 1)


def fill_product_row(ws, row, product, cfg):
    cartons = product.get("Cartons")
    unit_price = product.get("Unit Price (USD/KG)")

    ws.cell(row=row, column=2).value = product.get("Product Name") or ""
    ws.cell(row=row, column=4).value = specs_to_multiline(product.get("Specifications") or "")

    if cartons is not None and safe_str(cartons):
        try:
            ws.cell(row=row, column=6).value = int(float(safe_str(cartons)))
        except ValueError:
            ws.cell(row=row, column=6).value = cartons

    kg_per_carton = cfg.kg_per_carton
    raw_kg = product.get("KG/Carton")
    if raw_kg is not None and safe_str(raw_kg):
        try:
            kg_per_carton = float(safe_str(raw_kg))
        except ValueError:
            pass
    ws.cell(row=row, column=7).value = f"=F{row}*{kg_per_carton}"

    if unit_price is not None and safe_str(unit_price):
        try:
            ws.cell(row=row, column=8).value = float(safe_str(unit_price))
        except ValueError:
            ws.cell(row=row, column=8).value = unit_price
    ws.cell(row=row, column=9).value = f"=G{row}*H{row}"


def fill_header(ws, header, total_row, cfg):
    ws["B4"] = cfg.doc_title

    for cell, key in [("B7", "Seller Name"), ("B8", "Seller Address")]:
        if header.get(key):
            ws[cell] = header[key]

    inv = safe_str(header.get("Invoice No", ""))
    if inv:
        ws["F8"] = inv

    if header.get("Issue Date"):
        ws["H8"] = parse_date(header["Issue Date"])
    if header.get("Buyer Reference"):
        ws["F10"] = header["Buyer Reference"]
    if header.get("Due Date"):
        ws["H10"] = parse_date(header["Due Date"])
        ws["H10"].number_format = ws["H8"].number_format

    for cell, key in [
        ("B13", "Buyer Name"), ("B14", "Buyer Address"),
        ("F13", "Shipper Name"), ("F14", "Shipper Address"),
        ("B17", "Port of Loading"), ("D17", "Port of Destination"),
        ("B19", "Date of Loading"), ("D19", "Shipment Type"),
    ]:
        if header.get(key):
            ws[cell] = header[key]

    if header.get("Payment Terms"):
        ws["F17"] = " " + header["Payment Terms"]

    # Total formula
    _write_cell(ws, total_row, 8, f"=SUM(I{PRODUCT_FIRST_ROW}:I{total_row - 1})")

    # Incoterms / Currency (always total_row + 2)
    incoterms_row = total_row + 2
    _write_cell(ws, incoterms_row, 6,
                header.get("Incoterms") or ws.cell(row=incoterms_row, column=6).value)
    _write_cell(ws, incoterms_row, 8,
                header.get("Currency") or ws.cell(row=incoterms_row, column=8).value)

    shift = total_row - cfg.total_row_offset

    if header.get("Buyer Name"):
        _write_cell(ws, cfg.buyer_sig_row_base + shift, 6, header["Buyer Name"])

    if header.get("Bank Details"):
        _write_cell(ws, cfg.bank_details_row_base + shift, 2, "\n".join(
            p.strip() for p in header["Bank Details"].split("|")
        ))

    if header.get("Additional Info"):
        _write_cell(ws, total_row + 1, 2, header["Additional Info"])
